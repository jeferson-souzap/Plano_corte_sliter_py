"""
================================================================================
PLANO DE CORTE — OTIMIZADOR DE COMBINAÇÕES DE MATRIZES
================================================================================

OBJETIVO:
    Encontrar a melhor forma de combinar diferentes perfis (matrizes) em uma
    bobina de aço, minimizando o desperdício (refilo/perda).

REGRAS DE NEGÓCIO:
    1. Usuário escolhe: espessura + tipo de material + 1 a 3 matrizes âncora
    2. Ao menos 1 âncora deve ser selecionada (obrigatório)
    3. Todas as âncoras selecionadas DEVEM aparecer ao menos 1x na combinação
    4. Todas as âncoras devem ser da mesma espessura e tipo de material
    5. Matrizes NÃO selecionadas como âncora entram apenas como complementares
    6. Máximo de 1 matriz complementar por combinação
    7. Larguras de bobina fixas: 1200, 1000 ou 1500 mm
    8. Perda aceitável: entre 0,67% e 1,70% da largura
    9. Refilo mínimo por espessura:
       - Espessuras ≤ 3.0 mm: mínimo 10 mm de refilo
       - Espessuras > 3.0 mm: mínimo 14 mm de refilo
   10. Limite de cortes (opcional): máximo de cortes simultâneos

ESTRUTURA DO CÓDIGO:
    BLOCO 1: Configurações e constantes
    BLOCO 2: Funções de carga e limpeza de dados
    BLOCO 3: Funções de consulta ao banco de dados
    BLOCO 4: Motor de busca combinatorial (1, 2 ou 3 âncoras)
    BLOCO 5: Cálculo de KG
    BLOCO 6: Validação de resultados
    BLOCO 7: Interface com usuário (CLI)
    BLOCO 8: Exportação para Excel
    BLOCO 9: Função principal (main)
================================================================================
"""

import os
import platform
import pandas as pd
from itertools import combinations, product as iproduct
from datetime import datetime


# ================================================================================
# BLOCO 1: CONFIGURAÇÕES E CONSTANTES
# ================================================================================

SO      = platform.system()
USUARIO = os.getenv('USERNAME') if SO == 'Windows' else os.getenv('USER')

if SO == 'Windows':
    BASE_INPUT  = r'D:\#Mega\Jeferson - Dev\02 - Linguagens\Python\Acotel\Plano_corte_sliter_py\files\input'
    BASE_OUTPUT = r'D:\#Mega\Jeferson - Dev\02 - Linguagens\Python\Acotel\Plano_corte_sliter_py\files\output'
elif SO == 'Linux':
    BASE_INPUT  = r'/home/stark/Documentos/Dev/Plano_corte_sliter_py/files/input'
    BASE_OUTPUT = r'/home/stark/Documentos/Dev/Plano_corte_sliter_py/files/output'
else:
    raise Exception(f'Sistema operacional não suportado: {SO}')

# ── Parâmetros de negócio ──
LARGURAS_BOBINA      = [1200, 1000, 1500]
PERDA_MIN_PCT        = 0.67
PERDA_MAX_PCT        = 1.70
REFILO_MIN_ATE_3MM   = 10
REFILO_MIN_ACIMA_3MM = 14
MAX_ANCORAS          = 3    # máximo de âncoras selecionáveis pelo usuário
MAX_COMP_NA_COMBO    = 1    # máximo de matrizes complementares por combinação
PESO_MEDIO_BOB_PAD   = 12_000
QTD_BOBINAS_PAD      = 1


# ================================================================================
# BLOCO 2: FUNÇÕES DE CARGA E LIMPEZA DE DADOS
# ================================================================================

def carregar_dados(caminho: str) -> pd.DataFrame:
    """
    Carrega o arquivo Excel com as matrizes e faz limpeza dos dados.

    LIMPEZA REALIZADA:
        1. Remove espaços em branco das strings
        2. Converte espessura e desenvolvimento para números
        3. Remove linhas com dados ausentes ou inválidos
        4. Remove matrizes com desenvolvimento zero ou negativo
    """
    df = pd.read_excel(caminho)

    df['Código']           = df['Código'].astype(str).str.strip()
    df['Matriz']           = df['Matriz'].astype(str).str.strip()
    df['Tipo de material'] = df['Tipo de material'].astype(str).str.strip()

    df['Espessura']       = pd.to_numeric(df['Espessura'],       errors='coerce')
    df['Desenvolvimento'] = pd.to_numeric(df['Desenvolvimento'], errors='coerce')

    df = df.dropna(subset=['Espessura', 'Desenvolvimento', 'Matriz',
                            'Tipo de material', 'Código'])
    df = df[df['Desenvolvimento'] > 0]
    df = df[~df['Matriz'].isin(['nan', ''])]
    df = df[~df['Código'].isin(['nan', ''])]

    return df


# ================================================================================
# BLOCO 3: FUNÇÕES DE CONSULTA AO BANCO DE DADOS
# ================================================================================

def listar_espessuras(df: pd.DataFrame) -> list[float]:
    return sorted(df['Espessura'].unique())


def listar_tipos(df: pd.DataFrame, espessura: float) -> list[str]:
    return sorted(df[df['Espessura'] == espessura]['Tipo de material'].unique())


def listar_matrizes(df: pd.DataFrame, espessura: float, tipo: str) -> pd.DataFrame:
    """
    Retorna todas as matrizes disponíveis para espessura + tipo específicos.
    Quando uma matriz aparece múltiplas vezes, calcula a média do desenvolvimento.
    """
    mask = (df['Espessura'] == espessura) & (df['Tipo de material'] == tipo)
    return (
        df[mask]
        .groupby('Matriz')['Desenvolvimento']
        .mean()
        .reset_index()
        .rename(columns={'Desenvolvimento': 'Dev_mm'})
        .sort_values('Dev_mm', ascending=False)
        .reset_index(drop=True)
    )


def obter_desenvolvimento(df: pd.DataFrame, matriz: str, espessura: float) -> float:
    mask = (df['Matriz'] == matriz) & (df['Espessura'] == espessura)
    vals = df[mask]['Desenvolvimento'].dropna()
    if vals.empty:
        raise ValueError(f"Matriz '{matriz}' com espessura {espessura} mm não encontrada.")
    return float(vals.mean())


def obter_codigo(df: pd.DataFrame, matriz: str, espessura: float) -> str:
    mask = (df['Matriz'] == matriz) & (df['Espessura'] == espessura)
    vals = df[mask]['Código'].dropna()
    if vals.empty:
        return ''
    return str(vals.iloc[0])


# ================================================================================
# BLOCO 4: MOTOR DE BUSCA COMBINATÓRIA
# ================================================================================

def _montar_detalhe(matriz: str, codigo: str, dev: float, n_cortes: int) -> dict:
    """Cria um dict de detalhe padronizado para uma linha de resultado."""
    return {
        'Matriz':             matriz,
        'Codigo':             codigo,
        'Desenvolvimento_mm': dev,
        'N_cortes':           n_cortes,
        'Subtotal_mm':        round(dev * n_cortes, 3)
    }


def buscar_combinacoes_para_largura(
    df: pd.DataFrame,
    ancoras: list[str],
    devs_ancoras: list[float],
    codigos_ancoras: list[str],
    matrizes_comp: list[str],
    devs_comp: list[float],
    largura_bobina: int,
    espessura: float,
    limite_cortes: int | None = None
) -> list[dict]:
    """
    Motor principal: testa TODAS as combinações possíveis para uma largura,
    suportando 1, 2 ou 3 âncoras obrigatórias + até 1 complementar.

    ALGORITMO:
        Para cada tupla (n1, n2, ..., nK) de cortes de cada âncora:
            1. soma_ancoras = sum(dev_i * n_i)
            2. Se soma_ancoras > largura → descarta
            3. Valida combinação SEM complementar
            4. Para cada complementar que cabe no espaço restante:
               Valida combinação COM complementar (1x, 2x, ...)

    ENTRADA:
        df                : DataFrame completo (para busca de códigos de complementares)
        ancoras           : nomes das âncoras selecionadas (1 a 3)
        devs_ancoras      : desenvolvimentos em mm de cada âncora
        codigos_ancoras   : códigos dos itens de cada âncora
        matrizes_comp     : nomes das matrizes complementares candidatas
        devs_comp         : desenvolvimentos em mm de cada complementar
        largura_bobina    : 1000, 1200 ou 1500 mm
        espessura         : em mm (para calcular refilo mínimo)
        limite_cortes     : soma máxima de cortes (None = sem limite)

    SAÍDA:
        Lista de dicts, cada um representando uma combinação que entrou na
        janela de perda (0,67% – 1,70%), com status indicando se o refilo
        mínimo também foi respeitado.
    """
    perda_min_mm = largura_bobina * PERDA_MIN_PCT / 100
    perda_max_mm = largura_bobina * PERDA_MAX_PCT / 100
    refilo_min   = REFILO_MIN_ATE_3MM if espessura <= 3.0 else REFILO_MIN_ACIMA_3MM

    # Máximo de cortes que cada âncora pode ter individualmente na bobina
    max_cortes_ancoras = [max(1, int(largura_bobina / d)) for d in devs_ancoras]

    resultados = []

    # ── Itera sobre todas as combinações de quantidades de cortes das âncoras ──
    # Ex: 3 âncoras com max [4, 3, 2] → iproduct(range(1,5), range(1,4), range(1,3))
    for qtds_ancoras in iproduct(*[range(1, mx + 1) for mx in max_cortes_ancoras]):

        soma_ancoras = sum(d * n for d, n in zip(devs_ancoras, qtds_ancoras))

        # Descarta se a soma das âncoras já ultrapassou a largura
        if soma_ancoras > largura_bobina:
            continue

        espaco_restante      = largura_bobina - soma_ancoras
        total_cortes_ancoras = sum(qtds_ancoras)

        # ── Monta detalhes das âncoras (reutilizado nos dois casos abaixo) ──
        detalhes_ancoras = [
            _montar_detalhe(a, c, d, n)
            for a, c, d, n in zip(ancoras, codigos_ancoras, devs_ancoras, qtds_ancoras)
        ]

        ancora_str = ' + '.join(f'{a}(x{n})' for a, n in zip(ancoras, qtds_ancoras))

        # =================================================================
        # CASO 1: SÓ AS ÂNCORAS — sem complementar
        # =================================================================
        perda_mm     = espaco_restante
        total_cortes = total_cortes_ancoras
        passa_pct    = perda_min_mm <= perda_mm <= perda_max_mm
        passa_cortes = limite_cortes is None or total_cortes <= limite_cortes

        if passa_pct and passa_cortes:
            status = "✓ Válida" if perda_mm >= refilo_min else "Fora da regra"
            resultados.append({
                'Combinacao':     ancora_str,
                'Qtds_ancoras':   list(qtds_ancoras),
                'Num_comp':       0,
                'Total_cortes':   total_cortes,
                'Detalhes':       detalhes_ancoras[:],
                'Soma_cortes_mm': round(soma_ancoras, 3),
                'Perda_mm':       round(perda_mm, 3),
                'Perda_pct':      round(perda_mm / largura_bobina * 100, 4),
                'Largura_bobina': largura_bobina,
                'Status':         status
            })

        # =================================================================
        # CASO 2: ÂNCORAS + 1 COMPLEMENTAR
        # =================================================================
        # Filtra complementares que cabem no espaço restante
        indices_cabem = [
            i for i, dev in enumerate(devs_comp)
            if dev <= espaco_restante
        ]
        if not indices_cabem:
            continue

        for idx in indices_cabem:
            dev_c  = devs_comp[idx]
            nome_c = matrizes_comp[idx]
            max_nc = max(1, int(espaco_restante / dev_c))

            for n_c in range(1, max_nc + 1):
                soma_total   = soma_ancoras + dev_c * n_c
                total_cortes = total_cortes_ancoras + n_c

                if soma_total > largura_bobina:
                    break   # n_c só vai crescer, inutl continuar
                if limite_cortes is not None and total_cortes > limite_cortes:
                    continue

                perda_mm  = largura_bobina - soma_total
                passa_pct = perda_min_mm <= perda_mm <= perda_max_mm

                if passa_pct:
                    status   = "✓ Válida" if perda_mm >= refilo_min else "Fora da regra"
                    cod_c    = obter_codigo(df, nome_c, espessura)
                    detalhes = detalhes_ancoras + [
                        _montar_detalhe(nome_c, cod_c, dev_c, n_c)
                    ]
                    resultados.append({
                        'Combinacao':     f'{ancora_str} + {nome_c}(x{n_c})',
                        'Qtds_ancoras':   list(qtds_ancoras),
                        'Num_comp':       1,
                        'Total_cortes':   total_cortes,
                        'Detalhes':       detalhes,
                        'Soma_cortes_mm': round(soma_total, 3),
                        'Perda_mm':       round(perda_mm, 3),
                        'Perda_pct':      round(perda_mm / largura_bobina * 100, 4),
                        'Largura_bobina': largura_bobina,
                        'Status':         status
                    })

    return resultados


def encontrar_combinacoes(
    df: pd.DataFrame,
    espessura: float,
    tipo_material: str,
    ancoras: list[str],
    limite_cortes: int | None = None
) -> tuple[pd.DataFrame, int]:
    """
    Orquestrador: tenta larguras em sequência até encontrar resultados.

    ESTRATÉGIA:
        1. Tenta 1200 mm → 1000 mm → 1500 mm
        2. Para na primeira que retornar ao menos uma combinação

    VALIDAÇÕES ANTECIPADAS:
        - Cada âncora deve caber individualmente na largura
        - A soma mínima de todas as âncoras (1x cada) deve caber na largura
          (caso contrário, nunca haverá resultado para essa largura)
    """
    devs_ancoras    = [obter_desenvolvimento(df, a, espessura) for a in ancoras]
    codigos_ancoras = [obter_codigo(df, a, espessura)           for a in ancoras]
    soma_minima     = sum(devs_ancoras)   # 1 corte de cada âncora

    # Matrizes complementares: mesmo filtro, excluindo TODAS as âncoras
    mask_comp = (
        (df['Espessura'] == espessura) &
        (df['Tipo de material'] == tipo_material) &
        (~df['Matriz'].isin(ancoras))
    )
    candidatas = (
        df[mask_comp]
        .groupby('Matriz')['Desenvolvimento']
        .mean()
        .reset_index()
        .rename(columns={'Desenvolvimento': 'dev'})
        .sort_values('dev', ascending=False)
    )
    matrizes_comp = candidatas['Matriz'].tolist()
    devs_comp     = candidatas['dev'].tolist()

    for largura in LARGURAS_BOBINA:
        print(f"  → Tentando largura {largura} mm ...", end=' ')

        # Âncoras individuais cabem?
        nao_cabem = [a for a, d in zip(ancoras, devs_ancoras) if d > largura]
        if nao_cabem:
            print(f"âncora(s) {nao_cabem} não cabem individualmente. Pulando.")
            continue

        # Soma mínima cabe?
        if soma_minima > largura:
            print(f"soma mínima das âncoras ({soma_minima:.1f} mm) não cabe. Pulando.")
            continue

        resultados = buscar_combinacoes_para_largura(
            df=df,
            ancoras=ancoras,
            devs_ancoras=devs_ancoras,
            codigos_ancoras=codigos_ancoras,
            matrizes_comp=matrizes_comp,
            devs_comp=devs_comp,
            largura_bobina=largura,
            espessura=espessura,
            limite_cortes=limite_cortes
        )

        if resultados:
            print(f"{len(resultados)} combinações encontradas. ✓")
            df_res = (
                pd.DataFrame(resultados)
                .sort_values(['Perda_pct', 'Total_cortes', 'Num_comp'])
                .reset_index(drop=True)
            )
            return df_res, largura
        else:
            print("nenhuma combinação válida.")

    return pd.DataFrame(), 0


# ================================================================================
# BLOCO 5: CÁLCULO DE KG
# ================================================================================

def calcular_peso_medio_bobina(peso_informado: float, qtd_bobinas: int) -> float:
    """Peso médio por bobina = peso total ÷ quantidade de bobinas."""
    return peso_informado / qtd_bobinas


def calcular_kg_matriz(
    peso_medio_bobina: float,
    largura_bobina: int,
    n_cortes: int,
    desenvolvimento: float,
    qtd_bobinas: int
) -> float:
    """
    KG = (Peso_médio / Largura) × (N_cortes × Desenvolvimento × Qtd_bobinas)
    """
    return (peso_medio_bobina / largura_bobina) * (n_cortes * desenvolvimento * qtd_bobinas)


def calcular_kg_combinacao(
    detalhes: list[dict],
    peso_medio_bobina: float,
    largura_bobina: int,
    qtd_bobinas: int
) -> float:
    total = sum(
        calcular_kg_matriz(peso_medio_bobina, largura_bobina,
                           m['N_cortes'], m['Desenvolvimento_mm'], qtd_bobinas)
        for m in detalhes
    )
    return round(total, 2)


# ================================================================================
# BLOCO 6: VALIDAÇÃO DE RESULTADOS
# ================================================================================

def validar_resultado(df_res: pd.DataFrame, espessura: float) -> dict:
    refilo_min = REFILO_MIN_ATE_3MM if espessura <= 3.0 else REFILO_MIN_ACIMA_3MM
    validas    = len(df_res[df_res['Status'] == '✓ Válida'])
    fora_regra = len(df_res[df_res['Status'] == 'Fora da regra'])
    return {
        'total':      len(df_res),
        'validas':    validas,
        'fora_regra': fora_regra,
        'refilo_min': refilo_min
    }


# ================================================================================
# BLOCO 7: INTERFACE COM USUÁRIO (CLI)
# ================================================================================

def exibir_terminal(
    df_res: pd.DataFrame,
    largura: int,
    ancoras: list[str],
    espessura: float,
    tipo: str,
    limite_cortes: int | None = None
) -> None:
    """Exibe resultados formatados no terminal."""
    sep = "=" * 90
    print(f"\n{sep}")
    print("  PLANO DE CORTE — COMBINAÇÕES VÁLIDAS")
    print(sep)
    print(f"  Âncora(s)      : {' | '.join(ancoras)}")
    print(f"  Qtd. âncoras   : {len(ancoras)}")
    print(f"  Espessura      : {espessura} mm")
    print(f"  Tipo material  : {tipo}")
    print(f"  Largura bobina : {largura} mm")

    perda_min_mm = largura * PERDA_MIN_PCT / 100
    perda_max_mm = largura * PERDA_MAX_PCT / 100
    print(f"  Janela de perda: {PERDA_MIN_PCT}% – {PERDA_MAX_PCT}%  "
          f"|  {perda_min_mm:.2f} mm – {perda_max_mm:.2f} mm")

    refilo_min = REFILO_MIN_ATE_3MM if espessura <= 3.0 else REFILO_MIN_ACIMA_3MM
    simbolo    = '≤' if espessura <= 3.0 else '>'
    print(f"  Refilo mínimo  : {refilo_min} mm  (regra esp {simbolo} 3.0 mm)")

    if limite_cortes is not None:
        print(f"  Limite cortes  : {limite_cortes} cortes (soma total)")

    if df_res.empty:
        print(f"\n  ⚠  Nenhuma combinação válida encontrada.")
        print(f"     Sugestão: revise as âncoras ou amplie os parâmetros.")
        print(sep)
        return

    stats = validar_resultado(df_res, espessura)
    print(f"  Combinações    : {stats['total']} "
          f"({stats['validas']} válidas + {stats['fora_regra']} fora da regra)\n")

    fmt = "  {:<5} {:<54} {:<12} {:<12} {:<12} {}"
    print(fmt.format('#', 'Combinação', 'Soma (mm)', 'Perda (mm)', 'Perda (%)', 'Status'))
    print(fmt.format('-'*5, '-'*54, '-'*12, '-'*12, '-'*12, '-'*16))

    for i, r in df_res.iterrows():
        print(fmt.format(
            i + 1,
            r['Combinacao'][:53],
            f"{r['Soma_cortes_mm']:.2f}",
            f"{r['Perda_mm']:.3f}",
            f"{r['Perda_pct']:.4f}%",
            r['Status']
        ))

    print(sep)


def _selecionar_ancora(
    matrizes_disponiveis: pd.DataFrame,
    ja_selecionadas: list[str],
    numero_passo: int,
    ordinal: str
) -> str:
    """
    Exibe a lista de matrizes disponíveis (excluindo as já escolhidas)
    e retorna a âncora selecionada pelo usuário.
    """
    opcoes = matrizes_disponiveis[
        ~matrizes_disponiveis['Matriz'].isin(ja_selecionadas)
    ].reset_index(drop=True)

    if opcoes.empty:
        print(f"\n  ⚠ Não há matrizes disponíveis para a {ordinal} âncora.")
        return None

    print(f"\n[{numero_passo}] Escolha a {ordinal} ÂNCORA:")
    for i, row in opcoes.iterrows():
        print(f"    {i+1:3d}. {row['Matriz']:<38s}  dev = {row['Dev_mm']:.1f} mm")

    while True:
        try:
            escolha = int(input(f"\n  Número da {ordinal} âncora: "))
            return opcoes.iloc[escolha - 1]['Matriz']
        except (ValueError, IndexError):
            print("  ⚠ Inválido. Tente novamente.")


def menu_usuario(df: pd.DataFrame) -> tuple[float, str, list[str], int | None, int, float]:
    """
    Interface CLI: coleta todas as informações do usuário.

    PASSOS:
        [1]  Escolhe espessura
        [2]  Escolhe tipo de material
        [3]  Escolhe quantidade de âncoras (1, 2 ou 3)
        [4]  Escolhe 1ª âncora            (sempre)
        [5]  Escolhe 2ª âncora            (se qtd ≥ 2)
        [6]  Escolhe 3ª âncora            (se qtd = 3)
        [N]  Limite de cortes             (opcional)
        [N+1] Quantidade de bobinas
        [N+2] Peso total das bobinas

    SAÍDA:
        (espessura, tipo, lista_ancoras, limite_cortes, qtd_bobinas, peso_total)
    """
    print("\n" + "=" * 70)
    print("          SISTEMA DE PLANO DE CORTE")
    print(f"  Larguras testadas : {' → '.join(str(l) for l in LARGURAS_BOBINA)} mm")
    print(f"  Máximo de âncoras : {MAX_ANCORAS}")
    print(f"  Max complementares: {MAX_COMP_NA_COMBO} por combinação")
    print("=" * 70)

    # ── [1] Espessura ──
    espessuras = listar_espessuras(df)
    print("\n[1] Espessuras disponíveis:")
    for i, e in enumerate(espessuras, 1):
        print(f"    {i:3d}. {e} mm")

    while True:
        try:
            espessura = espessuras[int(input("\n  Número da espessura: ")) - 1]
            break
        except (ValueError, IndexError):
            print("  ⚠ Inválido. Tente novamente.")

    # ── [2] Tipo de material ──
    tipos = listar_tipos(df, espessura)
    print(f"\n[2] Tipos de material (esp={espessura} mm):")
    for i, t in enumerate(tipos, 1):
        print(f"    {i:3d}. {t}")

    while True:
        try:
            tipo = tipos[int(input("\n  Número do tipo: ")) - 1]
            break
        except (ValueError, IndexError):
            print("  ⚠ Inválido. Tente novamente.")

    # ── Lista completa de matrizes disponíveis para este filtro ──
    matrizes = listar_matrizes(df, espessura, tipo)
    total_matrizes = len(matrizes)

    # ── [3] Quantidade de âncoras ──
    max_possivel = min(MAX_ANCORAS, total_matrizes)
    print(f"\n[3] Quantidade de âncoras (1 a {max_possivel})")
    print(f"    Todas as âncoras escolhidas aparecerão obrigatoriamente")
    print(f"    em cada combinação gerada.")

    while True:
        try:
            qtd_ancoras = int(input(f"\n  Quantidade de âncoras [1]: ").strip() or "1")
            if not (1 <= qtd_ancoras <= max_possivel):
                raise ValueError
            break
        except ValueError:
            print(f"  ⚠ Digite um número entre 1 e {max_possivel}.")

    # ── [4], [5], [6] Seleção das âncoras ──
    ordinals = ["1ª", "2ª", "3ª"]
    ancoras  = []

    for i in range(qtd_ancoras):
        passo  = 4 + i
        ancora = _selecionar_ancora(matrizes, ancoras, passo, ordinals[i])
        if ancora is None:
            print(f"  Continuando com {len(ancoras)} âncora(s).")
            break
        ancoras.append(ancora)

    print(f"\n  Âncoras confirmadas ({len(ancoras)}):")
    for i, a in enumerate(ancoras, 1):
        dev = obter_desenvolvimento(df, a, espessura)
        print(f"    {i}. {a}  (dev = {dev:.1f} mm)")

    # ── Passo dinâmico: número do passo após as âncoras ──
    proximo_passo = 4 + len(ancoras)

    # ── Limite de cortes ──
    print(f"\n[{proximo_passo}] Limite de cortes por combinação (restrição de máquina)")
    print(f"    Deixe em branco para sem limite.")

    while True:
        entrada = input("  Limite de cortes: ").strip()
        if entrada == "":
            limite_cortes = None
            break
        try:
            limite_cortes = int(entrada)
            if limite_cortes < 1:
                raise ValueError
            break
        except ValueError:
            print("  ⚠ Digite um número positivo ou deixe em branco.")

    proximo_passo += 1

    # ── Quantidade de bobinas ──
    print(f"\n[{proximo_passo}] Quantidade de bobinas")
    while True:
        entrada = input(f"  Quantidade [{QTD_BOBINAS_PAD}]: ").strip()
        if entrada == "":
            qtd_bobinas = QTD_BOBINAS_PAD
            break
        try:
            qtd_bobinas = int(entrada)
            if qtd_bobinas < 1:
                raise ValueError
            break
        except ValueError:
            print("  ⚠ Digite um número positivo.")

    proximo_passo += 1

    # ── Peso total das bobinas ──
    print(f"\n[{proximo_passo}] Peso TOTAL do lote de bobinas (kg)")
    while True:
        entrada = input(f"  Peso total (kg) [{PESO_MEDIO_BOB_PAD:,.0f}]: ").strip()
        if entrada == "":
            peso_total = float(PESO_MEDIO_BOB_PAD)
            break
        try:
            # Aceita tanto vírgula quanto ponto como separador decimal
            peso_total = float(entrada.replace('.', '').replace(',', '.'))
            if peso_total <= 0:
                raise ValueError
            break
        except ValueError:
            print("  ⚠ Digite um número positivo (ex: 48000).")

    return espessura, tipo, ancoras, limite_cortes, qtd_bobinas, peso_total


# ================================================================================
# BLOCO 8: EXPORTAÇÃO PARA EXCEL
# ================================================================================

# Cores usadas nas abas — centralizadas para facilitar manutenção
_CORES = {
    'azul_esc':   "1F4E79",
    'azul_cla':   "BDD7EE",
    'verde':      "E2EFDA",
    'cinza':      "F2F2F2",
    'branco':     "FFFFFF",
    'ancora_1':   "FFF2CC",   # amarelo claro
    'ancora_2':   "FFE082",   # amarelo médio
    'ancora_3':   "FFB300",   # âmbar
    'complementar': "FFFFFF",
    'roxo':       "EDE7F6",
    'laranja':    "FFE0B2",
}

_COR_ANCORA_POR_INDICE = ['ancora_1', 'ancora_2', 'ancora_3']


def exportar_excel(
    df_res: pd.DataFrame,
    largura: int,
    ancoras: list[str],
    espessura: float,
    tipo: str,
    caminho: str,
    qtd_bobinas: int,
    peso_total: float,
    df_original: pd.DataFrame,
    limite_cortes: int | None = None
) -> None:
    """
    Exporta resultados para arquivo Excel com 2 abas.

    ABA 1 — Combinações:
        Cabeçalho com todos os parâmetros usados na busca.
        Tabela com: #, Combinação, Qtd Âncoras, Total Cortes,
                    Soma (mm), Perda (mm), Perda (%), KG, Status.

    ABA 2 — Detalhes:
        Uma linha por matriz por combinação.
        Papel: ÂNCORA 1 / ÂNCORA 2 / ÂNCORA 3 / Complementar
        (com cores diferentes por papel).
        Colunas: # Combo, Papel, Código, Matriz, Desenvolvimento (mm),
                 N° Cortes, Subtotal (mm), Qtd. KG.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb        = Workbook()
    ws_combos = wb.active
    ws_combos.title = "Combinações"
    ws_det    = wb.create_sheet("Detalhes")

    borda_fina = Side(style='thin', color='CCCCCC')
    borda      = Border(left=borda_fina, right=borda_fina,
                        top=borda_fina,  bottom=borda_fina)

    def cel(ws, linha, col, valor, negrito=False, fundo=_CORES['branco'],
            texto="000000", alinha="left", fmt=None, wrap=False):
        c = ws.cell(linha, col, valor)
        c.font      = Font(name="Arial", size=9, bold=negrito, color=texto)
        c.fill      = PatternFill("solid", start_color=fundo, end_color=fundo)
        c.alignment = Alignment(horizontal=alinha, vertical="center", wrap_text=wrap)
        c.border    = borda
        if fmt:
            c.number_format = fmt
        return c

    peso_medio = calcular_peso_medio_bobina(peso_total, qtd_bobinas)
    refilo_min = REFILO_MIN_ATE_3MM if espessura <= 3.0 else REFILO_MIN_ACIMA_3MM
    regra_ref  = (f"≤ 3.0 mm → {REFILO_MIN_ATE_3MM} mm  |  "
                  f"> 3.0 mm → {REFILO_MIN_ACIMA_3MM} mm")
    lim_str    = str(limite_cortes) if limite_cortes is not None else "Sem limite"

    # ════════════════════════════════════════════════════════════════
    # ABA 1: COMBINAÇÕES
    # ════════════════════════════════════════════════════════════════

    ws_combos.row_dimensions[1].height = 22
    ws_combos.cell(1, 1, "PLANO DE CORTE — COMBINAÇÕES VÁLIDAS").font = \
        Font(name="Arial", size=13, bold=True, color=_CORES['azul_esc'])
    ws_combos.merge_cells("A1:I1")

    # Bloco de parâmetros
    params = [("Qtd. de Âncoras", str(len(ancoras)))]
    for i, a in enumerate(ancoras, 1):
        dev = obter_desenvolvimento(df_original, a, espessura)
        params.append((f"  Âncora {i}", f"{a}  (dev = {dev:.1f} mm)"))
    params += [
        ("Espessura",         f"{espessura} mm"),
        ("Tipo de Material",  tipo),
        ("Largura da Bobina", f"{largura} mm"),
        ("Padrões Testados",  " → ".join(str(l) for l in LARGURAS_BOBINA)
                              + f"  (usado: {largura} mm)"),
        ("Limite de Cortes",  lim_str),
        ("Refilo Mínimo",     f"{refilo_min} mm  ({regra_ref})"),
        ("Qtd. de Bobinas",   str(qtd_bobinas)),
        ("Peso Total Lote",   f"{peso_total:,.0f} kg  ({peso_total/1000:.1f} ton)"),
        ("Peso Médio/Bobina", f"{peso_medio:,.0f} kg  ({peso_medio/1000:.2f} ton)"),
        ("Perda Mínima (%)",  f"{PERDA_MIN_PCT}%  ({largura * PERDA_MIN_PCT / 100:.2f} mm)"),
        ("Perda Máxima (%)",  f"{PERDA_MAX_PCT}%  ({largura * PERDA_MAX_PCT / 100:.2f} mm)"),
        ("Total Combinações", len(df_res)),
    ]

    for r, (chave, valor) in enumerate(params, start=2):
        cel(ws_combos, r, 1, chave, negrito=True, fundo=_CORES['azul_cla'])
        cel(ws_combos, r, 2, valor, fundo=_CORES['azul_cla'])
        ws_combos.merge_cells(f"B{r}:I{r}")

    lc = len(params) + 3   # linha do cabeçalho da tabela
    ws_combos.row_dimensions[lc].height = 28

    for col, titulo in enumerate(
        ["#", "Combinação", "Qtd Âncoras", "Total Cortes",
         "Soma (mm)", "Perda (mm)", "Perda (%)", "Qtd. KG", "Status"], 1
    ):
        cel(ws_combos, lc, col, titulo, negrito=True,
            fundo=_CORES['azul_esc'], texto=_CORES['branco'], alinha="center")

    for i, row in df_res.iterrows():
        ln      = lc + 1 + i
        zeb     = _CORES['verde'] if i % 2 == 0 else _CORES['cinza']
        kg      = calcular_kg_combinacao(row['Detalhes'], peso_medio, largura, qtd_bobinas)
        c_stat  = _CORES['verde'] if row['Status'] == "✓ Válida" else _CORES['laranja']
        qtd_a   = len(row['Qtds_ancoras'])

        cel(ws_combos, ln, 1, i + 1,                  fundo=zeb,               alinha="center")
        cel(ws_combos, ln, 2, row['Combinacao'],       fundo=zeb,               wrap=True)
        cel(ws_combos, ln, 3, qtd_a,                   fundo=_CORES['ancora_1'],alinha="center")
        cel(ws_combos, ln, 4, row['Total_cortes'],     fundo=zeb,               alinha="center")
        cel(ws_combos, ln, 5, row['Soma_cortes_mm'],   fundo=zeb,               alinha="right", fmt='#,##0.000')
        cel(ws_combos, ln, 6, row['Perda_mm'],         fundo=zeb,               alinha="right", fmt='#,##0.000')
        cel(ws_combos, ln, 7, row['Perda_pct'] / 100, fundo=zeb,               alinha="right", fmt='0.0000%')
        cel(ws_combos, ln, 8, kg,                      fundo=_CORES['roxo'],    alinha="right", fmt='#,##0.00')
        cel(ws_combos, ln, 9, row['Status'],           fundo=c_stat,            alinha="center")
        ws_combos.row_dimensions[ln].height = 16

    for col, w in zip("ABCDEFGHI", [5, 55, 12, 13, 16, 13, 11, 14, 13]):
        ws_combos.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════════
    # ABA 2: DETALHES
    # ════════════════════════════════════════════════════════════════

    ws_det.row_dimensions[1].height = 20
    ws_det.cell(1, 1, "DETALHES POR COMBINAÇÃO").font = \
        Font(name="Arial", size=12, bold=True, color=_CORES['azul_esc'])
    ws_det.merge_cells("A1:H1")

    for col, titulo in enumerate(
        ["# Combo", "Papel", "Código", "Matriz",
         "Desenvolvimento (mm)", "N° Cortes", "Subtotal (mm)", "Qtd. KG"], 1
    ):
        cel(ws_det, 2, col, titulo, negrito=True,
            fundo=_CORES['azul_esc'], texto=_CORES['branco'], alinha="center")

    ln = 3
    for i, row in df_res.iterrows():
        n_anc = len(row['Qtds_ancoras'])

        for j, det in enumerate(row['Detalhes']):
            # Define papel e cor da linha
            if j < n_anc:
                if n_anc == 1:
                    papel = "ÂNCORA"
                else:
                    papel = f"ÂNCORA {j + 1}"
                cor_linha = _CORES[_COR_ANCORA_POR_INDICE[j]]
                negrito_p = True
            else:
                papel     = "Complementar"
                cor_linha = _CORES['complementar']
                negrito_p = False

            kg_mat = calcular_kg_matriz(
                peso_medio, largura,
                det['N_cortes'], det['Desenvolvimento_mm'], qtd_bobinas
            )

            cel(ws_det, ln, 1, i + 1,                  alinha="center")
            cel(ws_det, ln, 2, papel,     fundo=cor_linha, alinha="center", negrito=negrito_p)
            cel(ws_det, ln, 3, det['Codigo'],           fundo=cor_linha)
            cel(ws_det, ln, 4, det['Matriz'],           fundo=cor_linha)
            cel(ws_det, ln, 5, det['Desenvolvimento_mm'], fundo=cor_linha, alinha="right", fmt='#,##0.000')
            cel(ws_det, ln, 6, det['N_cortes'],         fundo=cor_linha, alinha="center")
            cel(ws_det, ln, 7, det['Subtotal_mm'],      fundo=cor_linha, alinha="right", fmt='#,##0.000')
            cel(ws_det, ln, 8, kg_mat,                  fundo=_CORES['roxo'], alinha="right", fmt='#,##0.00')
            ln += 1

    for col, w in zip("ABCDEFGH", [10, 15, 16, 32, 22, 12, 16, 14]):
        ws_det.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(caminho) if os.path.dirname(caminho) else ".", exist_ok=True)
    wb.save(caminho)
    print(f"\n  Resultado exportado: {caminho}")


# ================================================================================
# BLOCO 9: FUNÇÃO PRINCIPAL (MAIN)
# ================================================================================

def main():
    """
    Fluxo principal:
        1. Carrega banco de dados
        2. Coleta parâmetros via menu (1 a 3 âncoras)
        3. Busca combinações válidas
        4. Exibe no terminal
        5. Exporta para Excel
    """
    caminho_db = os.path.join(BASE_INPUT, 'db_plano_corte.xlsx')
    print(f"\n  Carregando: {caminho_db}")
    df = carregar_dados(caminho_db)
    print(f"  {len(df)} produtos carregados.")

    espessura, tipo, ancoras, limite_cortes, qtd_bobinas, peso_total = menu_usuario(df)

    print(f"\n  Buscando combinações:")
    print(f"    Âncora(s) : {' | '.join(ancoras)}")
    print(f"    Espessura : {espessura} mm  |  Tipo: {tipo}")
    if limite_cortes:
        print(f"    Limite de cortes: {limite_cortes}")

    df_res, largura_usada = encontrar_combinacoes(
        df=df,
        espessura=espessura,
        tipo_material=tipo,
        ancoras=ancoras,
        limite_cortes=limite_cortes
    )

    exibir_terminal(
        df_res=df_res,
        largura=largura_usada,
        ancoras=ancoras,
        espessura=espessura,
        tipo=tipo,
        limite_cortes=limite_cortes
    )

    if not df_res.empty:
        # Nome do arquivo inclui qtd de âncoras + nomes sanitizados
        ancoras_safe = '_'.join(
            a.replace('/', '_').replace('"', 'in')
             .replace(',', '-').replace(' ', '_')
            for a in ancoras
        )
        ts       = datetime.now().strftime('%Y%m%d_%H%M%S')
        esp_str  = str(espessura).replace('.', '-')
        tipo_str = tipo.replace(' ', '_')
        qtda_str = f"{len(ancoras)}anc"

        nome = (f"plano_{qtda_str}_{ancoras_safe}"
                f"_esp{esp_str}_{tipo_str}_L{largura_usada}_{ts}.xlsx")
        caminho_out = os.path.join(BASE_OUTPUT, nome)

        exportar_excel(
            df_res=df_res,
            largura=largura_usada,
            ancoras=ancoras,
            espessura=espessura,
            tipo=tipo,
            caminho=caminho_out,
            qtd_bobinas=qtd_bobinas,
            peso_total=peso_total,
            df_original=df,
            limite_cortes=limite_cortes
        )


# ════════════════════════════════════════════════════════════════════════════════
# EXECUÇÃO
# ════════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print(f"  Usuário: {USUARIO}")
    main()